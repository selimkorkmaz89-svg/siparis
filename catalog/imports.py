"""Excel import for the product catalogue and the dealer account cards.

Rules from the specification:

* A code that appears more than once **in the uploaded file** aborts the whole
  import — nothing is written and the offending rows are listed.
* Every other row is classified as *new* or *update* and shown on a preview
  screen; the write only happens after the administrator confirms.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook, load_workbook

from catalog.models import DeviceModel, Product
from core.constants import Currency, ProductGroup
from dealers.models import Dealer

#: Column order matches Mikro's own stock export (KODU, ADI, FİYAT, DVZ,
#: TOPTAN KDV, MARKA, CİHAZ TÜRÜ, ÜRÜN GRUBU) so that file can be re-headered
#: and used almost as-is; tests_per_pack/description are ours, Mikro has no
#: equivalent column and they're left blank on that kind of import.
PRODUCT_COLUMNS = [
    ("code", _("Product code")),
    ("name", _("Product name")),
    ("price", _("List price")),
    ("price_currency", _("Currency")),
    ("vat_rate", _("VAT rate (%)")),
    ("brand", _("Brand")),
    ("device_model", _("Device model")),
    ("product_group", _("Product group")),
    ("tests_per_pack", _("Tests per pack")),
    ("description", _("Description")),
]

#: Labels for the fields actually written to Product - a superset of
#: PRODUCT_COLUMNS, since one input column (price + currency) becomes three
#: output fields (base_price_usd, list_price, price_currency).
PRODUCT_FIELD_LABELS = dict(PRODUCT_COLUMNS)
PRODUCT_FIELD_LABELS.update({
    "base_price_usd": _("List price (USD)"),
    "list_price": _("List price (native currency)"),
    "mikro_stok_kodu": _("Mikro stock code"),
})

#: Free-text currency names as they appear in a Mikro export, lower-cased.
#: "İ" lower-cases to "i̇" (with a combining dot) in Python, not plain "i" -
#: both the exact and the ASCII-ish spelling are listed so either survives a
#: copy/paste or a manual retype.
CURRENCY_TEXT_MAP = {
    "usd": Currency.USD,
    "amerikan doları": Currency.USD,
    "amerikan dolari": Currency.USD,
    "dolar": Currency.USD,
    "chf": Currency.CHF,
    "isviçre frangı": Currency.CHF,
    "i̇sviçre frangı": Currency.CHF,
    "isvicre frangi": Currency.CHF,
}

#: Free-text product group names as they appear in a Mikro export, with the
#: casing inconsistencies actually seen in practice ("Sarf" vs "sarf").
PRODUCT_GROUP_TEXT_MAP = {
    "cihaz": ProductGroup.DEVICE,
    "sarf": ProductGroup.CONSUMABLE,
    "yedek parça": ProductGroup.SPARE_PART,
    "yedek parca": ProductGroup.SPARE_PART,
}

DEALER_COLUMNS = [
    ("name", _("Dealer name")),
    ("code", _("Dealer code")),
    ("tax_no", _("Tax number")),
    ("tax_office", _("Tax office")),
    ("contact_person", _("Contact person")),
    ("phone", _("Phone")),
    ("email", _("Email")),
    ("city", _("City")),
    ("address", _("Address")),
    ("notes", _("Notes")),
]

SAMPLE_ROWS = {
    "product": [
        ["PRD-001", "Sample reagent kit", "125.00", "USD", "20", "Acme",
         "Acme X200", "Sarf", 100, ""]
    ],
    "dealer": [
        ["Sample Dealer Ltd.", "D-001", "1234567890", "Kadıköy", "Jane Doe",
         "+90 216 000 00 00", "info@sampledealer.com", "İstanbul", "", ""]
    ],
}


def column_spec(kind: str):
    return PRODUCT_COLUMNS if kind == "product" else DEALER_COLUMNS


def build_template(kind: str) -> bytes:
    columns = column_spec(kind)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "template"
    sheet.append([str(label) for _key, label in columns])
    for row in SAMPLE_ROWS[kind]:
        sheet.append(row)
    for index, (_key, label) in enumerate(columns, start=1):
        sheet.column_dimensions[chr(64 + index)].width = max(len(str(label)) + 6, 18)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@dataclass
class RowResult:
    row_no: int
    key: str
    action: str  # "create" | "update" | "error"
    values: dict = field(default_factory=dict)
    changes: list = field(default_factory=list)  # (label, old, new)
    error: str = ""


@dataclass
class ImportPreview:
    kind: str
    rows: list = field(default_factory=list)
    duplicates: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)

    @property
    def blocked(self) -> bool:
        """Duplicate keys or invalid rows stop the whole import."""
        return bool(self.duplicates or self.errors)

    @property
    def creates(self):
        return [row for row in self.rows if row.action == "create"]

    @property
    def updates(self):
        return [row for row in self.rows if row.action == "update"]


def _to_decimal(value, label) -> Decimal:
    if value in (None, ""):
        raise ValueError(_("%(field)s is required.") % {"field": label})
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, ValueError):
        raise ValueError(_("%(field)s must be a number.") % {"field": label})


def _to_int(value, label) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(Decimal(str(value).replace(",", ".").strip()))
    except (InvalidOperation, ValueError):
        raise ValueError(_("%(field)s must be a whole number.") % {"field": label})


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def parse_workbook(file_obj, kind: str) -> ImportPreview:
    """Validate an uploaded workbook and classify each row."""
    preview = ImportPreview(kind=kind)
    columns = column_spec(kind)
    try:
        workbook = load_workbook(file_obj, data_only=True)
    except Exception:
        preview.errors.append(_("The file could not be read. Please upload a valid .xlsx file."))
        return preview
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if not rows:
        preview.errors.append(_("The file contains no data rows."))
        return preview

    chf_to_usd_rate = None
    if kind == "product":
        from payments import services as fx

        current_rate = fx.get_rate()
        chf_to_usd_rate = current_rate.chf_to_usd_rate if current_rate else None

    diff_labels = PRODUCT_FIELD_LABELS if kind == "product" else dict(columns)
    seen: dict[str, int] = {}
    existing = _existing_keys(kind)
    chf_rows_without_rate = 0
    for offset, raw in enumerate(rows, start=2):
        values = list(raw) + [None] * (len(columns) - len(raw))
        if all(_text(value) == "" for value in values[: len(columns)]):
            continue
        data = {key: values[index] for index, (key, _label) in enumerate(columns)}
        key = _text(data[columns[0][0]])
        if kind == "product":
            key = key.upper()
        if not key:
            preview.errors.append(
                _("Row %(row)s: %(field)s is empty.")
                % {"row": offset, "field": columns[0][1]}
            )
            continue
        if key in seen:
            preview.duplicates.append(
                _("Row %(row)s: '%(key)s' already appears on row %(first)s.")
                % {"row": offset, "key": key, "first": seen[key]}
            )
            continue
        seen[key] = offset
        try:
            cleaned = _clean_row(data, kind, chf_to_usd_rate)
        except ValueError as exc:
            preview.errors.append(_("Row %(row)s: %(error)s") % {"row": offset, "error": exc})
            continue
        if kind == "product" and cleaned.get("price_currency") == Currency.CHF and not chf_to_usd_rate:
            chf_rows_without_rate += 1
        current = existing.get(key)
        if current is None:
            preview.rows.append(
                RowResult(row_no=offset, key=key, action="create", values=cleaned)
            )
        else:
            changes = _diff(current, cleaned, diff_labels)
            preview.rows.append(
                RowResult(
                    row_no=offset,
                    key=key,
                    action="update",
                    values=cleaned,
                    changes=changes,
                )
            )
    if not preview.rows and not preview.blocked:
        preview.errors.append(_("The file contains no importable rows."))
    if chf_rows_without_rate:
        preview.warnings.append(
            _(
                "%(count)s CHF-priced product(s) have no CHF/TRY rate to convert with yet "
                "and will import with a USD price of 0.00 - fetch the rate (System Settings) "
                "and they will be repriced automatically."
            )
            % {"count": chf_rows_without_rate}
        )
    return preview


def _map_choice(value, mapping: dict, label) -> str:
    key = _text(value).strip().lower()
    if not key:
        raise ValueError(_("%(field)s is required.") % {"field": label})
    if key not in mapping:
        raise ValueError(
            _("%(field)s '%(value)s' is not recognised.")
            % {"field": label, "value": _text(value)}
        )
    return mapping[key]


def _clean_row(data: dict, kind: str, chf_to_usd_rate: Decimal | None = None) -> dict:
    if kind == "product":
        price = _to_decimal(data["price"], _("List price"))
        currency = _map_choice(data["price_currency"], CURRENCY_TEXT_MAP, _("Currency"))
        if currency == Currency.CHF:
            list_price = price
            base_price_usd = (
                (price * chf_to_usd_rate).quantize(Decimal("0.01"))
                if chf_to_usd_rate else Decimal("0.00")
            )
        else:
            list_price = None
            base_price_usd = price
        code = _text(data["code"]).upper()
        return {
            "code": code,
            "name": _text(data["name"]),
            "brand": _text(data["brand"]),
            # Resolved to a DeviceModel instance in apply_preview - a plain
            # name here so blank cells and unrecognised names stay easy to spot
            # on the preview screen.
            "device_model": _text(data["device_model"]),
            "product_group": _map_choice(
                data["product_group"], PRODUCT_GROUP_TEXT_MAP, _("Product group")
            ),
            "tests_per_pack": _to_int(data["tests_per_pack"], _("Tests per pack")),
            "price_currency": currency,
            "list_price": list_price,
            "base_price_usd": base_price_usd,
            "vat_rate": _to_decimal(data["vat_rate"], _("VAT rate (%)")),
            "description": _text(data["description"]),
            # This importer's source is Mikro's own stock export, so the
            # product code already is the Mikro stock code.
            "mikro_stok_kodu": code,
        }
    return {
        "name": _text(data["name"]),
        "code": _text(data["code"]),
        "tax_no": _text(data["tax_no"]),
        "tax_office": _text(data["tax_office"]),
        "contact_person": _text(data["contact_person"]),
        "phone": _text(data["phone"]),
        "email": _text(data["email"]),
        "city": _text(data["city"]),
        "address": _text(data["address"]),
        "notes": _text(data["notes"]),
    }


def _existing_keys(kind: str) -> dict:
    if kind == "product":
        return {
            product.code.upper(): {
                "code": product.code,
                "name": product.name,
                "brand": product.brand,
                "device_model": product.device_model.name if product.device_model_id else "",
                "product_group": product.product_group,
                "tests_per_pack": product.tests_per_pack,
                "price_currency": product.price_currency,
                "list_price": product.list_price,
                "base_price_usd": product.base_price_usd,
                "vat_rate": product.vat_rate,
                "description": product.description,
                "mikro_stok_kodu": product.mikro_stok_kodu,
            }
            for product in Product.objects.select_related("device_model")
        }
    return {
        dealer.name: {
            "name": dealer.name,
            "code": dealer.code,
            "tax_no": dealer.tax_no,
            "tax_office": dealer.tax_office,
            "contact_person": dealer.contact_person,
            "phone": dealer.phone,
            "email": dealer.email,
            "city": dealer.city,
            "address": dealer.address,
            "notes": dealer.notes,
        }
        for dealer in Dealer.objects.all()
    }


def _same(old, new) -> bool:
    """Compare two cell values, treating 20 and 20.00 as equal."""
    if isinstance(old, (Decimal, int, float)) or isinstance(new, (Decimal, int, float)):
        try:
            return Decimal(str(old or 0)) == Decimal(str(new or 0))
        except InvalidOperation:
            pass
    return str(old or "") == str(new or "")


def _diff(current: dict, new: dict, labels: dict) -> list:
    changes = []
    for key, value in new.items():
        old = current.get(key)
        if not _same(old, value):
            changes.append((labels.get(key, key), old, value))
    return changes


def apply_preview(preview: ImportPreview) -> dict:
    """Write a validated preview to the database."""
    created = updated = 0
    model = Product if preview.kind == "product" else Dealer
    lookup = "code" if preview.kind == "product" else "name"
    for row in preview.rows:
        values = dict(row.values)
        # device_model isn't a plain Product field - it's a name string that
        # resolves to a DeviceModel row, so it is set after create/update.
        device_model_name = values.pop("device_model", None) if model is Product else None
        if row.action == "create":
            obj = model.objects.create(**values)
            created += 1
        elif row.action == "update":
            filter_value = values[lookup]
            obj = model.objects.filter(**{f"{lookup}__iexact": filter_value}).first()
            if obj is None:
                obj = model.objects.create(**values)
                created += 1
            else:
                for key, value in values.items():
                    setattr(obj, key, value)
                obj.save()
                updated += 1
        if model is Product:
            _apply_device_model(obj, device_model_name, values.get("brand", ""))
    return {"created": created, "updated": updated}


def _apply_device_model(product: Product, name: str | None, brand: str) -> None:
    """Resolve an imported device-model name to a row, creating it if new."""
    name = (name or "").strip()
    if not name:
        if product.device_model_id:
            product.device_model = None
            product.save(update_fields=["device_model"])
        return
    device_model, _created = DeviceModel.objects.get_or_create(
        name=name, defaults={"brand": brand}
    )
    if product.device_model_id != device_model.id:
        product.device_model = device_model
        product.save(update_fields=["device_model"])


def preview_to_session(preview: ImportPreview) -> dict:
    return {
        "kind": preview.kind,
        "rows": [
            {
                "row_no": row.row_no,
                "key": row.key,
                "action": row.action,
                "values": {
                    k: (None if v is None else str(v)) for k, v in row.values.items()
                },
            }
            for row in preview.rows
        ],
    }


def preview_from_session(payload: dict) -> ImportPreview:
    preview = ImportPreview(kind=payload["kind"])
    for row in payload["rows"]:
        values = row["values"]
        if payload["kind"] == "product":
            values = {
                **values,
                "tests_per_pack": int(values["tests_per_pack"]),
                "list_price": Decimal(values["list_price"]) if values.get("list_price") else None,
                "base_price_usd": Decimal(values["base_price_usd"]),
                "vat_rate": Decimal(values["vat_rate"]),
            }
        preview.rows.append(
            RowResult(
                row_no=row["row_no"], key=row["key"], action=row["action"], values=values
            )
        )
    return preview
