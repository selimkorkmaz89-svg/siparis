"""Excel export helpers shared by every list screen."""
from __future__ import annotations

import datetime as dt
from decimal import Decimal
from typing import Iterable, Sequence

from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="0D8DBE")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _clean(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dt.datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.replace(tzinfo=None)
    if value is None:
        return ""
    if isinstance(value, (str, int, float, dt.date)):
        return value
    return str(value)


def build_workbook(title: str, headers: Sequence[str], rows: Iterable[Sequence]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(title)[:31] or "Sheet1"
    sheet.append([str(h) for h in headers])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append([_clean(value) for value in row])
    widths = [len(str(h)) + 4 for h in headers]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for index, value in enumerate(row):
            if index < len(widths):
                widths[index] = max(widths[index], min(len(str(value)) + 3, 60))
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    return workbook


def excel_response(filename: str, title: str, headers, rows) -> HttpResponse:
    workbook = build_workbook(title, headers, rows)
    stamp = timezone.localdate().strftime("%Y%m%d")
    safe_name = slugify(filename) or "export"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{safe_name}-{stamp}.xlsx"'
    workbook.save(response)
    return response
